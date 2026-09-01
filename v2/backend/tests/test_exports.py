"""
Tests for exports:
1. WhatsApp 6-line format (asserts no markdown chars *, _, #, `).
2. HTML daily report snapshot against reference.
3. Excel operations export column & structure verification.
"""
from __future__ import annotations

import datetime
from decimal import Decimal
import io
import openpyxl
import pytest

from app.exports.daily_html import render_daily_report_html
from app.exports.whatsapp import build_whatsapp_text
from app.exports.xlsx import export_operations_to_xlsx


def test_whatsapp_no_markdown_chars():
    text = build_whatsapp_text(
        date=datetime.date(2026, 7, 31),
        unit_code="ATM",
        actual_eff=Decimal("89.48"),
        target_eff=Decimal("93.00"),
        rupee_lost=Decimal("486000"),
        top_issue_label="Electrical",
        top_issue_duration_str="3h 12m",
        top_issue_pct=Decimal("42"),
        best_loom_no="AJ-017",
        best_loom_eff=Decimal("96.1"),
        worst_loom_no="SZ-004",
        worst_loom_eff=Decimal("74.8"),
    )

    lines = text.strip().split("\n")
    assert len(lines) == 6, f"Expected 6 lines, got {len(lines)}"

    # Ensure no markdown styling chars
    for bad in ["*", "_", "#", "`"]:
        assert bad not in text, f"Found banned markdown character '{bad}' in WhatsApp output"

    assert "ATM Weaving - 31 Jul 2026" in lines[0]
    assert "Efficiency: 89.5% (target 93.0%)" in lines[1]
    assert "Loss: Rs.4,86,000 (EST)" in lines[2]
    assert "Top issue: Electrical - 3h 12m (42%)" in lines[3]


def test_html_daily_report_snapshot():
    html = render_daily_report_html(
        date=datetime.date(2026, 7, 31),
        unit_code="ATM",
        shifts_summary=[
            {"shift_code": "1", "target_eff": 93.0, "actual_eff": 89.48, "metres": 11150, "kilo_picks": 5000, "stopped_minutes_total": 45},
            {"shift_code": "2", "target_eff": 93.0, "actual_eff": 91.18, "metres": 11361, "kilo_picks": 5100, "stopped_minutes_total": 30},
            {"shift_code": "3", "target_eff": 93.0, "actual_eff": 88.03, "metres": 11408, "kilo_picks": 5050, "stopped_minutes_total": 60},
        ],
        day_total={"actual_eff": 89.56, "metres": 33919, "kilo_picks": 15150, "stopped_minutes_total": 135},
        top_issues=[
            {"reason_label_en": "Electrical stop", "total_minutes": 180, "pct_of_loom_downtime": 40.0, "count": 3}
        ],
    )

    assert "<!DOCTYPE html>" in html
    assert "Daily Weaving Production Report" in html
    assert "@media print" in html
    assert "window.print()" in html
    assert "Shift 1" in html
    assert "33,919" in html
    assert "Electrical stop" in html


def test_xlsx_operations_export_structure():
    sample_rows = [
        {
            "loom_no": "AJ-001",
            "loom_type_code": "810",
            "shed_code": "SHED-1",
            "style_code": "STYLE-A",
            "shift_code": "1",
            "weaver_name": "Muthu Kumar",
            "metres": 120.5,
            "kilo_picks": 45.2,
            "loom_efficiency_pct": 91.2,
            "performance_eff_pct": 95.0,
            "utilization_pct": 96.0,
            "cohort_gap_pp": 1.5,
            "warp_breaks_per_1000": 0.05,
            "weft_breaks_per_1000": 0.10,
            "stopped_minutes": 20,
            "rupee_lost": {"value": 1500, "rate_source": "ESTIMATED"},
            "status": "GREEN",
        }
    ]

    xlsx_bytes = export_operations_to_xlsx(sample_rows, title="ATM_Test")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["ATM_Test"]

    assert ws.cell(row=1, column=1).value == "Loom No"
    assert ws.cell(row=1, column=9).value == "Loom Eff %"
    assert ws.cell(row=2, column=1).value == "AJ-001"
    assert ws.cell(row=2, column=9).value == 91.2
