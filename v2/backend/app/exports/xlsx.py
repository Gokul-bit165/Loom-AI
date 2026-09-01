"""
Loom AI v2 — Exports: Excel (.xlsx) table exporter.
"""
from __future__ import annotations

import io
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_operations_to_xlsx(rows: list[dict[str, Any]], title: str = "Operations") -> bytes:
    """
    Generates styled Excel table matching Loom AI design tokens.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    headers = [
        "Loom No", "Type", "Shed", "Style", "Shift", "Weaver",
        "Metres", "Kilo-Picks", "Loom Eff %", "Perf Eff %", "Util %",
        "vs Cohort (pp)", "Warp Br/1000", "Weft Br/1000", "Stopped (min)", "Loss (Rs)", "Status"
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for r in rows:
        rupee_val = r.get("rupee_lost")
        if isinstance(rupee_val, dict):
            rupee_num = rupee_val.get("value")
        elif hasattr(rupee_val, "value"):
            rupee_num = rupee_val.value
        else:
            rupee_num = rupee_val

        ws.append([
            r.get("loom_no"),
            r.get("loom_type_code"),
            r.get("shed_code") or "-",
            r.get("style_code"),
            r.get("shift_code"),
            r.get("weaver_name") or "-",
            float(r.get("metres") or 0),
            float(r.get("kilo_picks") or 0),
            float(r.get("loom_efficiency_pct") or 0) if r.get("loom_efficiency_pct") is not None else None,
            float(r.get("performance_eff_pct") or 0) if r.get("performance_eff_pct") is not None else None,
            float(r.get("utilization_pct") or 0) if r.get("utilization_pct") is not None else None,
            float(r.get("cohort_gap_pp") or 0) if r.get("cohort_gap_pp") is not None else None,
            float(r.get("warp_breaks_per_1000") or 0) if r.get("warp_breaks_per_1000") is not None else None,
            float(r.get("weft_breaks_per_1000") or 0) if r.get("weft_breaks_per_1000") is not None else None,
            r.get("stopped_minutes", 0),
            float(rupee_num) if rupee_num is not None else 0,
            r.get("status", "GREY"),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = right_align

    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
