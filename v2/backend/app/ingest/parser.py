"""
Loom AI v2 — Ingest parser for real mill sheets.

Supports:
1. MILL_DAILY_PREP_WIDE:
   - Metric-as-rows, shifts-as-columns (I, II, III, TOTAL, REMARKS).
   - Extracts available hours from "TARGET @ X HRS" labels (e.g. 22.5, 22.0, 21.0).
   - Remarks column parsed into shift-associated remarks.
   - Extracts stated EFFI % for downstream cross-check verification.

2. LOOM_LONG_FORMAT:
   - Standard columnar CSV/Excel with per-loom rows.
"""
from __future__ import annotations

import csv
import datetime
import io
import re
from typing import Any, Optional
from decimal import Decimal

import openpyxl

from app.ingest.registry import Template, MILL_DAILY_PREP_WIDE, LOOM_LONG_FORMAT


class HeaderMappingError(Exception):
    pass


def _extract_available_hours(label: str) -> Optional[Decimal]:
    """Extracts available hours from labels like 'TARGET @ 22.5 HRS' or 'TARGET @ 22 HRS'."""
    m = re.search(r"@\s*(\d+(?:\.\d+)?)\s*HRS?", label, re.IGNORECASE)
    if m:
        return Decimal(m.group(1))
    return None


def parse_wide_prep_sheet(
    file_bytes: bytes,
    work_date: datetime.date,
) -> list[dict[str, Any]]:
    """
    Parses ATM's Daily Preparatory Production Report (Excel format).
    Transforms wide shift columns (I, II, III) into normalized shift records.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HeaderMappingError("Empty workbook.")

    # Find header row containing SHIFT, I, II, III
    header_idx = -1
    for idx, r in enumerate(rows):
        r_str = [str(c).strip().upper() if c is not None else "" for c in r]
        if "SHIFT" in r_str or "I" in r_str:
            header_idx = idx
            break

    if header_idx == -1:
        # Fallback to row 0
        header_idx = 0

    headers = [str(c).strip().upper() if c is not None else "" for c in rows[header_idx]]
    
    # Map shift columns
    shift_cols = {}
    remarks_col_idx = None
    metric_col_idx = 0

    for idx, h in enumerate(headers):
        if h in ("I", "1", "SHIFT 1", "SHIFT I"):
            shift_cols["1"] = idx
        elif h in ("II", "2", "SHIFT 2", "SHIFT II"):
            shift_cols["2"] = idx
        elif h in ("III", "3", "SHIFT 3", "SHIFT III"):
            shift_cols["3"] = idx
        elif "REMARK" in h:
            remarks_col_idx = idx

    if not shift_cols:
        # Default standard positions if not explicitly found
        shift_cols = {"1": 1, "2": 2, "3": 3}
        if len(headers) > 5:
            remarks_col_idx = 5

    # Collect remarks lines from the sheet
    shift_remarks: dict[str, list[str]] = {"1": [], "2": [], "3": []}
    current_shift_for_remarks = "1"
    
    # Read rows beneath header
    data_by_metric: dict[str, dict[str, Any]] = {}
    available_hours_found: Optional[Decimal] = None

    for row_num, r in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if not r or all(c is None for c in r):
            continue
        
        metric_label = str(r[metric_col_idx]).strip() if r[metric_col_idx] is not None else ""
        if not metric_label:
            continue

        # Check for remarks in remarks column
        if remarks_col_idx is not None and remarks_col_idx < len(r) and r[remarks_col_idx]:
            rem_text = str(r[remarks_col_idx]).strip()
            # Detect shift section heading in remarks (e.g. "SHIFT : 1", "SHIFT: 2")
            shift_match = re.search(r"SHIFT\s*:\s*([123I]+)", rem_text, re.IGNORECASE)
            if shift_match:
                s_val = shift_match.group(1).upper()
                current_shift_for_remarks = "1" if s_val in ("1", "I") else ("2" if s_val in ("2", "II") else "3")
            else:
                shift_remarks[current_shift_for_remarks].append(rem_text)

        # Check for TARGET @ X HRS
        hrs = _extract_available_hours(metric_label)
        if hrs:
            available_hours_found = hrs

        # Normalize metric category
        norm_label = metric_label.upper()
        if "ACTUAL" in norm_label:
            key = "ACTUAL"
        elif "TARGET" in norm_label:
            key = "TARGET"
        elif "SCHEDULED" in norm_label or "24 HRS" in norm_label:
            key = "SCHEDULED"
        elif "EFFI" in norm_label or "EFF%" in norm_label:
            key = "STATED_EFFI_PCT"
        else:
            key = norm_label

        data_by_metric[key] = {
            "label": metric_label,
            "values": {
                shift_code: r[col_idx] if col_idx < len(r) else None
                for shift_code, col_idx in shift_cols.items()
            }
        }

    # Transform wide data into normalized list of shift records
    records: list[dict[str, Any]] = []
    for shift_code in ("1", "2", "3"):
        actual_val = data_by_metric.get("ACTUAL", {}).get("values", {}).get(shift_code)
        target_val = data_by_metric.get("TARGET", {}).get("values", {}).get(shift_code)
        sched_val = data_by_metric.get("SCHEDULED", {}).get("values", {}).get(shift_code)
        stated_effi = data_by_metric.get("STATED_EFFI_PCT", {}).get("values", {}).get(shift_code)
        
        remarks_joined = "; ".join(shift_remarks.get(shift_code, [])) or None

        records.append({
            "_row_index": header_idx + 1,
            "_work_date": work_date,
            "SHIFT": shift_code,
            "ACTUAL_PRODUCTION": actual_val,
            "TARGET_PRODUCTION": target_val,
            "SCHEDULED_PRODUCTION": sched_val,
            "STATED_EFFI_PCT": stated_effi,
            "AVAILABLE_HOURS": available_hours_found or Decimal("22.5"),
            "REMARKS": remarks_joined,
        })

    return records


def parse_long_loom_sheet(
    file_bytes: bytes,
    work_date: datetime.date,
    is_csv: bool = False,
) -> list[dict[str, Any]]:
    """Parses clean per-loom layout."""
    if is_csv:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        result = []
        for idx, row in enumerate(reader, start=2):
            if all(not v.strip() for v in row.values() if v):
                continue
            rec = {"_row_index": idx, "_work_date": work_date}
            for k, v in row.items():
                if k:
                    rec[k.strip().upper()] = v.strip() if v else None
            result.append(rec)
        return result

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().upper() if h is not None else "" for h in next(rows_iter)]
    result = []
    for idx, row in enumerate(rows_iter, start=2):
        if all(v is None for v in row):
            continue
        rec = {"_row_index": idx, "_work_date": work_date}
        for h, val in zip(headers, row):
            if h:
                rec[h] = val
        result.append(rec)
    return result
